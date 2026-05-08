"""
Normalization helpers and shared constants for DocketPro CRM web.
"""

from datetime import date, datetime
from typing import Optional

# --- String normalization (mirrors dashboard_app.py norm()) ---

_NORM = {
    "active":    "Активний",
    "inactive":  "Неактивний",
    "paid":      "Оплачено",
    "unpaid":    "Не оплачено",
    "cancelled": "Скасовано",
    "canceled":  "Скасовано",
    "draft":     "Чернетка",
    "signed":    "Підписано",
    "pending":   "На підписанні",
    "access":    "Доступ",
    "hourly":    "Погодинний",
}


def norm(value: Optional[str]) -> str:
    """Normalize English legacy values to Ukrainian canonical strings."""
    if value is None:
        return ""
    s = str(value).strip()
    return _NORM.get(s.lower(), s)


# --- Date helpers ---

DATE_FMT = "%d.%m.%Y"


def parse_date(val) -> Optional[date]:
    """Parse date from Excel cell value (datetime, date, or string dd.mm.yyyy)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def fmt_date(d: Optional[date]) -> str:
    """Format date as dd.mm.yyyy string, or empty string if None."""
    if d is None:
        return ""
    return d.strftime(DATE_FMT)


def fmt_money(val) -> str:
    """Format float as Ukrainian currency string."""
    if val is None:
        return "0,00"
    try:
        return f"{float(val):,.2f}".replace(",", " ").replace(".", ",")
    except (ValueError, TypeError):
        return "0,00"


# --- Invoice/act status helpers ---

def is_paid(pay_status: str) -> bool:
    n = norm(pay_status)
    return n == "Оплачено"


def is_cancelled(status: str) -> bool:
    n = norm(status)
    return n == "Скасовано"


def is_overdue(pay_status: str, due_date_str: str) -> bool:
    """True if invoice is unpaid and due date has passed."""
    if is_paid(pay_status) or is_cancelled(pay_status):
        return False
    d = parse_date(due_date_str)
    if d is None:
        return False
    return d < date.today()


# --- PDF URL helper ---

def pdf_url(abs_path: str) -> str:
    """Convert absolute pdf_path from DB to a /docs/... web URL."""
    if not abs_path:
        return ""
    try:
        from config import CONTRACTS_DIR as _base
    except Exception:
        _base = ""
    rel = abs_path
    if _base:
        base = _base.rstrip("/\\")
        if rel.startswith(base):
            rel = rel[len(base):].lstrip("/\\")
    return "/docs/" + rel.replace("\\", "/")


# --- Excel export helper ---

def make_xlsx(headers: list, rows: list, sheet_name: str = "Дані") -> "io.BytesIO":
    """
    Build an xlsx workbook from headers (list[str]) and rows (list[list]).
    Returns BytesIO ready for StreamingResponse.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="F1F5F9")
    THIN      = Side(style="thin", color="CBD5E1")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT      = Alignment(horizontal="left",   vertical="center")

    # Header row
    ws.row_dimensions[1].height = 22
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 18
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border    = BORDER
            cell.alignment = LEFT
            cell.font      = Font(size=10)

    # Auto column widths (capped at 50)
    for col_idx, h in enumerate(headers, 1):
        col_vals = [str(h)] + [str(r[col_idx - 1]) if r[col_idx - 1] is not None else ""
                                for r in rows]
        width = min(max(len(v) for v in col_vals) + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --- Categories ---

EXPENSE_CATEGORIES = [
    "Дизайн",
    "Юридичні витрати",
    "Технічна підтримка",
    "Розробка програмного забезпечення",
    "Реклама та маркетинг",
]

APP_PAYMENT_SOURCES = ["lyqpay", "Приват24", "Банк", "Готівка", "Інше"]
