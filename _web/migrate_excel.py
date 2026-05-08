"""
migrate_excel.py — One-time import from CRM_DOCKETPRO_2026.xlsx → crm.db (SQLite).

Usage:
    python3 _web/migrate_excel.py
    # or from _web/ directory:
    python3 migrate_excel.py
"""

import sys
import sqlite3
from pathlib import Path
from datetime import date, datetime

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl

CRM_ROOT   = Path(__file__).parent.parent
EXCEL_PATH = CRM_ROOT / "CRM_DOCKETPRO_2026.xlsx"
DB_PATH    = CRM_ROOT / "crm.db"

# ── helpers ──────────────────────────────────────────────────────────────────

_NORM = {
    "active":    "Активний",   "inactive":  "Неактивний",
    "paid":      "Оплачено",   "unpaid":    "Не оплачено",
    "cancelled": "Скасовано",  "canceled":  "Скасовано",
    "draft":     "Чернетка",   "signed":    "Підписано",
    "pending":   "На підписанні",
    "access":    "Доступ",     "hourly":    "Погодинний",
}

def norm(v):
    if not v or not isinstance(v, str):
        return v or ""
    return _NORM.get(v.strip().lower(), v.strip())

def parse_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, date):
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return s

def flt(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default

def nt(val) -> str:
    return str(val).strip() if val is not None else ""

def cidx(headers, name, default):
    try:
        return headers.index(name)
    except ValueError:
        return default

# ── migration ─────────────────────────────────────────────────────────────────

def migrate():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    from database import init_db
    init_db()
    print(f"Database: {DB_PATH}")

    # Single raw connection; FK OFF to avoid ordering constraints
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys=OFF")
    counts = {}

    # ── Clients ───────────────────────────────────────────────────────────────
    if "Clients" in wb.sheetnames:
        ws   = wb["Clients"]
        ch   = [c.value for c in ws[1]]
        n    = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            edrpou  = nt(row[cidx(ch,"EDRPOU",0)])
            name    = nt(row[cidx(ch,"ClientName",1)])
            director= nt(row[cidx(ch,"Director",2)])
            email   = nt(row[cidx(ch,"Email",3)])
            phone   = nt(row[cidx(ch,"Phone",4)])
            address = nt(row[cidx(ch,"Address",5)])
            status  = norm(row[cidx(ch,"Status",6)]) or "Активний"
            if not edrpou or not name:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO clients "
                "(edrpou,name,director,email,phone,address,status) VALUES (?,?,?,?,?,?,?)",
                (edrpou, name, director, email, phone, address, status)
            )
            n += 1
        conn.commit()
        counts["clients"] = n
    else:
        counts["clients"] = 0

    # ── Contracts ─────────────────────────────────────────────────────────────
    ws  = wb["Contracts"]
    ch  = [c.value for c in ws[1]]
    n   = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cno     = nt(row[cidx(ch,"ContractNo",0)])
        edrpou  = nt(row[cidx(ch,"EDRPOU",1)])
        client  = nt(row[cidx(ch,"ClientName",2)])
        cdate   = parse_date(row[cidx(ch,"ContractDate",3)])
        cend    = parse_date(row[cidx(ch,"ContractEnd",4)])
        curr    = nt(row[cidx(ch,"Currency",5)]) or "UAH"
        tariff  = flt(row[cidx(ch,"TariffFX",6)])
        tyrate  = nt(row[cidx(ch,"TypeRate",7)])
        users   = int(flt(row[cidx(ch,"Users",8)]) or 1)
        status  = norm(row[cidx(ch,"Status",9)]) or "Активний"
        ctype   = norm(row[cidx(ch,"ContractType",10)]) or "Доступ"
        subject = nt(row[cidx(ch,"Subject",11)])
        hrate   = flt(row[cidx(ch,"HourRate",12)])
        pdfpath = nt(row[cidx(ch,"PdfPath",13)])
        nbu     = flt(row[cidx(ch,"NbuRate",14)])
        acttempl= nt(row[cidx(ch,"ActTemplate",15)]) if len(row) > 15 else ""
        if not cno:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO contracts "
            "(contract_no,edrpou,client_name,contract_date,contract_end,"
            "currency,tariff_fx,type_rate,users,status,contract_type,"
            "subject,hour_rate,pdf_path,nbu_rate,act_template) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (cno, edrpou, client, cdate, cend, curr, tariff, tyrate,
             users, status, ctype, subject, hrate, pdfpath, nbu, acttempl)
        )
        n += 1
    conn.commit()
    counts["contracts"] = n

    # ── Invoices ──────────────────────────────────────────────────────────────
    ws  = wb["Invoices"]
    ih  = [c.value for c in ws[1]]
    n   = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        inv_no   = nt(row[cidx(ih,"InvoiceNo",0)])
        cno      = nt(row[cidx(ih,"ContractNo",1)])
        client   = nt(row[cidx(ih,"ClientName",2)])
        idate    = parse_date(row[cidx(ih,"InvoiceDate",3)])
        fxrate   = flt(row[cidx(ih,"FxRate",4)])
        curr     = nt(row[cidx(ih,"Currency",5)])
        sumfx    = flt(row[cidx(ih,"SumFX",6)])
        sumuah   = flt(row[cidx(ih,"SumUAH",7)])
        pfrom    = parse_date(row[cidx(ih,"PeriodFrom",8)])
        pto      = parse_date(row[cidx(ih,"PeriodTo",9)])
        due      = parse_date(row[cidx(ih,"DueDate",10)])
        pstatus  = norm(row[cidx(ih,"PayStatus",11)]) or "Не оплачено"
        paydate  = parse_date(row[cidx(ih,"PayDate",12)])
        itype    = norm(row[cidx(ih,"InvoiceType",13)]) if len(row) > 13 else ""
        months   = int(flt(row[cidx(ih,"Months",14)]) or 1) if len(row) > 14 else 1
        sumwords = nt(row[cidx(ih,"SumWords",15)]) if len(row) > 15 else ""
        pdfpath  = nt(row[cidx(ih,"PdfPath",16)]) if len(row) > 16 else ""
        disc     = flt(row[cidx(ih,"DiscountPct",17)]) if len(row) > 17 else 0
        if not inv_no:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO invoices "
            "(invoice_no,contract_no,client_name,invoice_date,fx_rate,currency,"
            "sum_fx,sum_uah,period_from,period_to,due_date,pay_status,pay_date,"
            "invoice_type,months,sum_words,pdf_path,discount_pct) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (inv_no, cno, client, idate, fxrate or None, curr,
             sumfx, sumuah, pfrom, pto, due, pstatus, paydate or None,
             itype, months, sumwords, pdfpath, disc)
        )
        n += 1
    conn.commit()
    counts["invoices"] = n

    # ── Acts ──────────────────────────────────────────────────────────────────
    if "Acts" in wb.sheetnames:
        ws  = wb["Acts"]
        ah  = [c.value for c in ws[1]]
        n   = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            act_no  = nt(row[cidx(ah,"ActNo",0)])
            inv_no  = nt(row[cidx(ah,"InvoiceNo",1)]) if len(row) > 1 else ""
            cno     = nt(row[cidx(ah,"ContractNo",2)]) if len(row) > 2 else ""
            client  = nt(row[cidx(ah,"ClientName",3)]) if len(row) > 3 else ""
            adate   = parse_date(row[cidx(ah,"ActDate",4)]) if len(row) > 4 else ""
            sumuah  = flt(row[cidx(ah,"SumUAH",5)]) if len(row) > 5 else 0
            pfrom   = parse_date(row[cidx(ah,"PeriodFrom",6)]) if len(row) > 6 else ""
            pto     = parse_date(row[cidx(ah,"PeriodTo",7)]) if len(row) > 7 else ""
            status  = norm(row[cidx(ah,"Status",8)]) if len(row) > 8 else "Чернетка"
            pdfpath = nt(row[cidx(ah,"PdfPath",9)]) if len(row) > 9 else ""
            if not act_no:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO acts "
                "(act_no,invoice_no,contract_no,client_name,act_date,"
                "period_from,period_to,sum_uah,status,pdf_path) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (act_no, inv_no or None, cno, client, adate,
                 pfrom, pto, sumuah, status or "Чернетка", pdfpath)
            )
            n += 1
        conn.commit()
        counts["acts"] = n
    else:
        counts["acts"] = 0

    # ── Expenses ──────────────────────────────────────────────────────────────
    if "Expenses" in wb.sheetnames:
        ws  = wb["Expenses"]
        n   = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            edate   = parse_date(row[0])
            cat     = nt(row[1])
            desc    = nt(row[2])
            amount  = flt(row[3])
            curr    = nt(row[4]) or "грн"
            exrate  = flt(row[5], 1)
            amuah   = flt(row[6]) if len(row) > 6 and row[6] else amount
            conn.execute(
                "INSERT INTO expenses "
                "(exp_date,category,description,amount,currency,exchange_rate,amount_uah) "
                "VALUES (?,?,?,?,?,?,?)",
                (edate, cat, desc, amount, curr, exrate, amuah)
            )
            n += 1
        conn.commit()
        counts["expenses"] = n
    else:
        counts["expenses"] = 0

    # ── AppPayments ───────────────────────────────────────────────────────────
    if "AppPayments" in wb.sheetnames:
        ws  = wb["AppPayments"]
        n   = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            pdate  = parse_date(row[0])
            amount = flt(row[1])
            curr   = nt(row[2]) or "UAH"
            desc   = nt(row[3]) if len(row) > 3 else ""
            source = nt(row[4]) if len(row) > 4 else "lyqpay"
            exists = conn.execute(
                "SELECT id FROM app_payments WHERE pay_date=? AND amount=?",
                (pdate, amount)
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO app_payments (pay_date,amount,currency,description,source) "
                    "VALUES (?,?,?,?,?)",
                    (pdate, amount, curr, desc, source)
                )
                n += 1
        conn.commit()
        counts["app_payments"] = n
    else:
        counts["app_payments"] = 0

    conn.close()

    print("\n✅ Migration complete:")
    for k, v in counts.items():
        print(f"   {k:15s}: {v} rows")


if __name__ == "__main__":
    migrate()
